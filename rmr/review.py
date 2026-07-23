"""Step 5: export blind human-review spreadsheets (one .xlsx per screening step).

The design is a census with adjudication. Every item the LLM screened at steps 2, 3 and 4 is
listed with three empty columns: ``Review 1``, where a first reviewer judges EVERY row;
``Review 2``, where a second reviewer adjudicates only the rows on which reviewer 1 differed
from the model; and ``Finale``, the consolidated verdict. Both work BLIND by hiding the
columns to their left, so neither is anchored by the model (nor, for the adjudicator, by
reviewer 1). The model's own verdict rides along in ``llm_decision`` purely as a reference,
and the authoritative copy stays in the step JSON files, joined back by ``id``.

Each sheet aggregates all origins into one workbook; the ``id`` keeps its origin prefix
(SC / GO / GH / HF), so the origin is always recoverable. This reads only the step outputs and
the captured-content files, so it needs no LLM call.

Sheets (``data/review/<MODEL_EXPERIMENT>/``), each ending in the three review columns:
- ``step-2-review.xlsx``: id, title, llm_decision.
- ``step-3-review.xlsx``: id, abstract, keywords, llm_decision. For grey literature the
  ``abstract`` cell composes the abstract/summary with its publication date and link.
- ``step-4-answers-review.xlsx``: the research answers (solution name, RQa-RQh, IC2, IC3) and
  llm_decision. Step 4 has a SINGLE sheet: the answers carry the evidence the reviewer needs,
  so the verdict is recorded on that same row.

Reviewers save their filled copy as ``step-<n>-review-reviewed.xlsx``, which steps 6 to 8 read;
re-running this step therefore never overwrites their work.
"""

import json
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from rmr import config
from rmr.paths import abstract_path, ensure_parent, review_dir, step_output_path

ORIGINS = ["scopus", "google", "github", "hf"]
# Origins whose step-3 content is an LLM summary of a web page: for these the abstract cell
# composes the summary with its publication date and link. scopus and hf carry a real abstract
# (Scopus gateway / arXiv), so they go in as-is. Mirrors GREY_SUMMARY_ORIGINS in screening.
COMPOSE_ORIGINS = {"google", "github"}
DECISION_VALUES = '"Include,Exclude"'  # Excel list-validation formula for the review columns
AGREEMENT_VALUES = '"Yes,No"'          # ... for the step-4 `Agreement` column
AGREEMENT_COL = "Agreement"
# The three columns the reviewers fill. Reviewer 1 screens EVERY row; reviewer 2 adjudicates
# only the rows where reviewer 1 differs from the model; `Finale` carries the consolidated
# verdict that step 8 applies back to the pipeline. Both reviewers work blind, hiding the
# columns to their left (`llm_decision`, and `Review 1` for the adjudicator).
R1_COL, R2_COL, FINAL_COL = "Review 1", "Review 2", "Finale"
REVIEW_COLS = [R1_COL, R2_COL, FINAL_COL]
RQ_KEYS = ["RQa", "RQb", "RQc", "RQd", "RQe", "RQf", "RQg", "RQh"]
# The step-4 sheet: the research answers give the reviewer the evidence, and the verdict is
# recorded on the same row. `Agreement` (Yes/No) audits the extracted answers themselves,
# separately from the include/exclude verdict in the three review columns.
ANSWERS_HEADER = (["ID", "Title", "Abstract", "Solution name"] + RQ_KEYS
                  + ["IC2", "IC3", "llm_decision", AGREEMENT_COL] + REVIEW_COLS)
ANSWERS_DROPDOWNS = {AGREEMENT_COL: AGREEMENT_VALUES,
                     **{name: DECISION_VALUES for name in REVIEW_COLS}}


def _records(origin: str, step: int) -> list[dict] | None:
    """The step's screened records for one origin, or None when the step has not run there."""
    path = step_output_path(origin, step)
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8")).get("records", [])


def _captured(origin: str, item_id: str) -> dict:
    """The captured/summarized content for one item (content/abstract/<id>.json), or {}."""
    path = abstract_path(origin, item_id)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def _compose_abstract(origin: str, record: dict, captured: dict) -> str:
    """The abstract cell for step 3.

    - No content (e.g. EC3, content unavailable for scraping): fall back to the title and link
      so the cell is never blank and the reviewer can still identify/locate the source.
    - scopus and hf: the real abstract, used as-is.
    - google/github (COMPOSE_ORIGINS): the LLM summary composed with its publication date and
      link, since that summary may drop those (per the review spec).
    """
    data = record.get("data", {})
    link = data.get("link") or captured.get("link") or ""
    text = captured.get("abstract") or captured.get("summary") or ""
    if not text.strip():
        title = data.get("title") or captured.get("title") or ""
        return "\n".join([f"Title: {title}", f"Link: {link}"])
    if origin not in COMPOSE_ORIGINS:
        return text
    return "\n".join([
        text,
        f"Publication date: {captured.get('publication_date', '')}",
        f"Link: {link}",
    ])


def _step2_rows() -> list[list]:
    rows = []
    for origin in ORIGINS:
        records = _records(origin, 2)
        if records is None:
            print(f"[review] step 2: no output for '{origin}', skipped")
            continue
        for record in records:
            rows.append([record.get("id", ""), record.get("title", ""), ""])
    return rows


def _step3_rows() -> list[list]:
    rows = []
    for origin in ORIGINS:
        records = _records(origin, 3)
        if records is None:
            print(f"[review] step 3: no output for '{origin}', skipped")
            continue
        for record in records:
            rid = record.get("id", "")
            captured = _captured(origin, rid)
            keywords = ", ".join(captured.get("keywords", []))
            rows.append([rid, _compose_abstract(origin, record, captured), keywords, ""])
    return rows


def _rq_cell(answers: dict, key: str) -> str:
    """One research-question answer as ``note`` / ``value`` / ``evidence`` lines. RQf's value is
    a list; it is joined into one line."""
    rq = answers.get(key) or {}
    value = rq.get("value", "")
    if isinstance(value, list):
        value = ", ".join(value)
    return "\n".join([
        f"note: {rq.get('note', '')}",
        f"value: {value}",
        f"evidence: {rq.get('evidence', '')}",
    ])


def _ic2_cell(criteria: dict) -> str:
    ic2 = criteria.get("IC2") or {}
    return "\n".join([
        f"type: {ic2.get('type', '')}",
        f"activity: {ic2.get('software_engineering_activity', '')}",
        f"note: {ic2.get('note', '')}",
    ])


def _ic3_cell(criteria: dict) -> str:
    ic3 = criteria.get("IC3") or {}
    return "\n".join([
        f"type: {ic3.get('type', '')}",
        f"note: {ic3.get('note', '')}",
    ])


def _step4_answers_rows() -> list[list]:
    """One row per step-4 record: the extracted research answers and the IC2/IC3
    classification. Records with no answers (e.g. EC3, content unavailable) still get a row,
    with the answer cells left empty, so the sheet covers the whole step."""
    rows = []
    for origin in ORIGINS:
        records = _records(origin, 4)
        if records is None:
            print(f"[review] step 4 answers: no output for '{origin}', skipped")
            continue
        for record in records:
            answers = record.get("answers") or {}
            rid = record.get("id", "")
            title = record.get("data", {}).get("title", "")
            abstract = _captured(origin, rid).get("abstract") or _captured(origin, rid).get("summary") or ""
            criteria = (record.get("screening") or {}).get("criteria") or {}
            row = [rid, title, abstract, answers.get("solution_name", "")]
            row += [_rq_cell(answers, key) if answers else "" for key in RQ_KEYS]
            row += [_ic2_cell(criteria), _ic3_cell(criteria), ""]
            rows.append(row)
    return rows


def _answers_review_rows() -> list[list]:
    """The step-4 review row: the research answers, then the model's final verdict (the
    reference column the reviewer hides), the empty ``Agreement`` cell and the three empty
    review columns."""
    llm = _llm_decisions(4)
    return [row[:-1] + [llm.get(str(row[0]), "").capitalize(), "", "", "", ""]
            for row in _step4_answers_rows()]


# Column widths per header name (Excel character units); wrapped columns hold long text.
_WIDTHS = {
    "id": 10, "ID": 10, "title": 60, "Title": 60, "abstract": 90, "Abstract": 80,
    "keywords": 40, "file": 18, "llm_decision": 14, "Agreement": 12, "Review 1": 12, "Review 2": 12,
    "Finale": 12, "Solution name": 24, "IC2": 40, "IC3": 40,
}
_WRAP = {"title", "Title", "abstract", "Abstract", "keywords", "IC2", "IC3"}


def _col_width(name: str) -> int:
    return 50 if name.startswith("RQ") else _WIDTHS.get(name, 20)


def _col_wrap(name: str) -> bool:
    return name.startswith("RQ") or name in _WRAP


def _drive_folder_url() -> str:
    """The shared Google Drive folder holding the step-4 source files, from
    ``REVIEW_DRIVE_FOLDER_URL``. Empty when unset (file cells then stay plain text)."""
    config.load_dotenv()
    return os.environ.get("REVIEW_DRIVE_FOLDER_URL", "").strip()


def _write_sheet(path, header: list[str], rows: list[list], dropdowns: dict | None = None,
                 link_column: str | None = None, link_url: str = "") -> None:
    """Write one review workbook: a bold, frozen header, the data rows, and a list-validation
    dropdown on each column named in ``dropdowns`` ({header name: Excel list formula}).

    When ``link_column`` and ``link_url`` are given, every non-empty cell in that column
    becomes a hyperlink to ``link_url`` (the shared Drive folder), so the reviewer opens the
    folder in one click and finds the file by the name shown in the cell.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append(row)
    for col, name in enumerate(header, start=1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = _col_width(name)
        if _col_wrap(name):
            for cell in ws[letter][1:]:  # data rows only
                cell.alignment = wrap
        if name == link_column and link_url:
            for cell in ws[letter][1:]:
                if cell.value:  # keep empty cells (unavailable file) as-is
                    cell.hyperlink = link_url
                    cell.style = "Hyperlink"

    for name, formula in (dropdowns or {}).items():
        if name not in header:
            continue
        letter = get_column_letter(header.index(name) + 1)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(validation)
        if rows:
            validation.add(f"{letter}2:{letter}{len(rows) + 1}")

    ensure_parent(path)
    wb.save(path)


def _screening_review_rows(rows: list[list], step: int) -> list[list]:
    """Build a screening review row: the LLM decision (capitalized, the reference column the
    reviewer hides) followed by the three empty review columns.
    Layout: [id, content..., llm_decision, Review 1, Review 2, Finale]. Steps 6-8 read those
    columns by name, not by position."""
    llm = _llm_decisions(step)
    return [row[:-1] + [llm.get(str(row[0]), "").capitalize(), "", "", ""] for row in rows]


def export_review_sheets() -> dict:
    """Generate the blind human-review spreadsheets for the current experiment."""
    out = review_dir()
    folder_url = _drive_folder_url()
    if not folder_url:
        print("[review] REVIEW_DRIVE_FOLDER_URL not set: step-4 file names stay plain text "
              "(set it to the shared Drive folder link to make them clickable)")
    # Every sheet ends with the three review columns (Review 1 / Review 2 / Finale) and carries
    # the `llm_decision` reference column that the reviewer hides. Steps 6-8 read those columns
    # by name, so column position does not matter and the reviewer may hide or reorder freely.
    decision_dd = {name: DECISION_VALUES for name in REVIEW_COLS}
    # (name, header, rows, dropdowns, link_column) — link_column is hyperlinked to folder_url.
    sheets = [
        ("step-2-review.xlsx", ["id", "title", "llm_decision"] + REVIEW_COLS,
         _screening_review_rows(_step2_rows(), 2), decision_dd, None),
        ("step-3-review.xlsx", ["id", "abstract", "keywords", "llm_decision"] + REVIEW_COLS,
         _screening_review_rows(_step3_rows(), 3), decision_dd, None),
        ("step-4-answers-review.xlsx", ANSWERS_HEADER, _answers_review_rows(),
         ANSWERS_DROPDOWNS, None),
    ]
    counts = {}
    for name, header, rows, dropdowns, link_column in sheets:
        path = out / name
        _write_sheet(path, header, rows, dropdowns, link_column=link_column, link_url=folder_url)
        counts[name] = len(rows)
        print(f"[review] {name}: {len(rows)} rows -> {path}")
    return counts


# --- Step 6: report the rows awaiting adjudication ----------------------------------------
# Reviewer 2 adjudicates exactly the rows where reviewer 1 differs from the model. Both
# reviewers work in the SAME workbook, so this step builds no sheet: it lists which rows need
# ``Review 2`` and which of those are still open.

INCLUDE_EXCLUDE = ("include", "exclude")
YES_NO = ("yes", "no")

# The reviewers' filled copy of each sheet, read by steps 6 to 8. Keeping it separate from the
# exported ``step-<n>-review.xlsx`` means re-running step 5 never overwrites their work.
# Step 4 has a SINGLE sheet: the research answers carry enough of the full text for the
# reviewer to judge, so the verdict is given on that same sheet.
REVIEWED_SHEETS = {
    2: "step-2-review-reviewed.xlsx",
    3: "step-3-review-reviewed.xlsx",
    4: "step-4-answers-review-reviewed.xlsx",
}

# A step's full sheet is usually already under review by the time the previous step's human
# verdict admits new records (its residuals), so those cannot simply be appended. They go to a
# small supplementary sheet instead, read side by side with the main one.
RESIDUAL_SHEETS = {3: "step-3-residuals-review.xlsx",
                   4: "step-4-answers-residuals-review.xlsx"}
RESIDUAL_REVIEWED_SHEETS = {step: name.replace(".xlsx", "-reviewed.xlsx")
                            for step, name in RESIDUAL_SHEETS.items()}


def reviewed_column(review, step: int, header: str) -> dict | None:
    """``{id: value}`` for one column of a step's reviewed sheet, merged with its residuals
    sheet. Reviewers may keep the residuals apart or paste them into the main sheet; either way
    the merged view is the same. Returns None when neither sheet exists."""
    main = _sheet_column(review / REVIEWED_SHEETS[step], header)
    extra = (_sheet_column(review / RESIDUAL_REVIEWED_SHEETS[step], header)
             if step in RESIDUAL_REVIEWED_SHEETS else None)
    if main is None and extra is None:
        return None
    merged = dict(main or {})
    merged.update(extra or {})
    return merged


def _llm_decisions(step: int) -> dict:
    """{id: 'include'|'exclude'} decided by the LLM at ``step`` (2/3/4), across all origins."""
    out = {}
    for origin in ORIGINS:
        records = _records(origin, step)
        if records is None:
            continue
        for record in records:
            rid = record.get("id")
            decision = (record.get("decision") if step == 2
                        else (record.get("screening") or {}).get("decision"))
            if rid and decision:
                out[str(rid)] = decision.strip().lower()
    return out


def _sheet_column(path, header_name: str) -> dict | None:
    """{id: value} for the column headed ``header_name`` in a filled sheet, lower-cased.
    None when the sheet does not exist; {} when the column is absent; '' for a blank cell.

    The header is matched case-insensitively and ignoring surrounding spaces, so a reviewer's
    copy that types ``review 1`` instead of ``Review 1`` still lines up.
    """
    if not path.exists():
        return None
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    wanted = header_name.strip().lower()
    col = next((i + 1 for i, name in enumerate(headers)
                if isinstance(name, str) and name.strip().lower() == wanted), None)
    if col is None:
        return {}
    out = {}
    for row in range(2, ws.max_row + 1):
        rid = ws.cell(row, 1).value
        if rid is None:
            continue
        value = ws.cell(row, col).value
        out[str(rid)] = str(value).strip().lower() if value is not None else ""
    return out


def _needs_adjudication(step: int, r1: dict) -> tuple[set, int]:
    """Return (ids needing ``Review 2``, count of rows reviewer 1 has not judged yet).

    A row needs adjudication when reviewer 1's Include/Exclude differs from the model's.
    """
    llm = _llm_decisions(step)
    conflicts, blanks = set(), 0
    for rid, value in r1.items():
        if value not in INCLUDE_EXCLUDE:
            blanks += 1
            continue
        model = llm.get(rid)
        if model and model != value:
            conflicts.add(rid)
    return conflicts, blanks


def first_pass_column(review, step: int) -> dict | None:
    """``{id: verdict}`` with the FIRST-PASS human verdict of a step.

    Normally that is ``Review 1``. On some rows the verdict was typed one column across, into
    ``Review 2``, leaving ``Review 1`` blank; since ``Review 2`` only ever holds an adjudication
    when the row is a conflict, a blank ``Review 1`` with a filled ``Review 2`` is read as the
    first pass. Rows resolved this way agree with the model by construction, so no adjudication
    is looked up for them afterwards.
    """
    r1 = reviewed_column(review, step, R1_COL)
    if r1 is None:
        return None
    r2 = reviewed_column(review, step, R2_COL) or {}
    return {rid: (value or r2.get(rid, "")) for rid, value in r1.items()}


def _residual_ids(step: int) -> set:
    """Ids that entered ``step`` through the previous step's human review (its residuals)."""
    path = review_dir() / "residuals" / f"step-{step - 1}-residuals.json"
    if not path.exists():
        return set()
    data = json.loads(path.read_text(encoding="utf-8"))
    return {str(item["id"]) for item in data.get("residuals", [])}


def export_residual_sheets() -> dict:
    """Export a review sheet holding ONLY the records that entered each step as residuals.

    Same columns as the step's full sheet, so the reviewers judge the newcomers alongside the
    review already in progress instead of having to reopen it.
    """
    out = review_dir()
    folder_url = _drive_folder_url()
    specs = [
        (3, ["id", "abstract", "keywords", "llm_decision"] + REVIEW_COLS,
         lambda: _screening_review_rows(_step3_rows(), 3),
         {name: DECISION_VALUES for name in REVIEW_COLS}),
        (4, ANSWERS_HEADER, _answers_review_rows, ANSWERS_DROPDOWNS),
    ]
    counts = {}
    for step, header, builder, dropdowns in specs:
        ids = _residual_ids(step)
        if not ids:
            print(f"[residuals] step {step}: no residuals recorded, skipped")
            continue
        rows = [row for row in builder() if str(row[0]) in ids]
        if not rows:
            print(f"[residuals] step {step}: the {len(ids)} residual id(s) are not in the "
                  f"step-{step} output yet; run step {step} for their origins first")
            continue
        path = out / RESIDUAL_SHEETS[step]
        _write_sheet(path, header, rows, dropdowns, link_url=folder_url)
        counts[RESIDUAL_SHEETS[step]] = len(rows)
        print(f"[residuals] {RESIDUAL_SHEETS[step]}: {len(rows)} of {len(ids)} residual(s) "
              f"-> {path}")

    if not counts:
        raise RuntimeError(
            "no residual sheet to export; run step 8 to record the residuals first"
        )
    return counts


def report_tiebreaks() -> dict:
    """Step 6: list the rows that need adjudication (``Review 2``) in each reviewed sheet.

    Reviewer 2 fills ``Review 2`` only where reviewer 1 differs from the model, so this tells
    the adjudicator exactly which rows to open, flags the ones still pending, and warns about
    rows adjudicated without an underlying conflict.
    """
    review = review_dir()
    specs = [(f"step-{step}", REVIEWED_SHEETS[step], step) for step in sorted(REVIEWED_SHEETS)]

    report = {}
    for label, filename, step in specs:
        # merges the step's sheet with its residuals sheet
        r1 = first_pass_column(review, step)
        r2 = reviewed_column(review, step, R2_COL) or {}
        if r1 is None:
            print(f"[tiebreak] {label}: {filename} not found, skipped")
            continue
        if not r1:
            print(f"[tiebreak] {label}: {filename} has no '{R1_COL}' column, skipped")
            continue
        expected = INCLUDE_EXCLUDE
        conflicts, blanks = _needs_adjudication(step, r1)
        pending = sorted(rid for rid in conflicts if r2.get(rid, "") not in expected)
        spurious = sorted(rid for rid, value in r2.items()
                          if value in expected and rid not in conflicts)
        report[label] = {"sheet": filename, "rows": len(r1), "not_reviewed": blanks,
                         "conflicts": len(conflicts), "pending": pending,
                         "adjudicated_without_conflict": spurious}
        print(f"[tiebreak] {label}: {len(conflicts)} of {len(r1)} rows need '{R2_COL}' "
              f"({blanks} not reviewed by reviewer 1, {len(pending)} still pending)")
        if pending:
            shown = ", ".join(pending[:20]) + (" ..." if len(pending) > 20 else "")
            print(f"[tiebreak] {label} pending -> {shown}")
        if spurious:
            print(f"[tiebreak] {label} warning: {len(spurious)} row(s) have '{R2_COL}' filled "
                  f"without a conflict ({', '.join(spurious[:5])})")

    if not report:
        raise RuntimeError(
            f"no reviewed sheet found in {review}; expected one of: "
            + ", ".join(REVIEWED_SHEETS.values())
        )
    out = review / "tiebreak" / "pending.json"
    ensure_parent(out)
    out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"[tiebreak] wrote {out}")
    return report
